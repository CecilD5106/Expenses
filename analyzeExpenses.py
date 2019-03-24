import openpyxl
from datetime import date
from datetime import datetime

# Program variable
categories = []
filePath = 'E:\\Excel\\'
# Calculate expenses per category
def calculateExpenses(myWS, myES, start, end):
    # Loop through the expenses in Expense Sheet (ES)
    for i in range(3, 50):
        # Initialize expense to 0
        exp = 0
        # Get expense category from ES
        cat = myES.cell(row=i, column=1).value
        # Loop through worksheet to get expenses
        # for a date range
        for j in range(3, 10001):
            getDate = myWS.cell(row=j, column=1).value
            if getDate != None:
                dt = getDate.date()
                # Verify if in date range
                if start <= dt <= end:
                    # Verify if in category
                    if cat == myWS.cell(row=j, column=10).value:
                        # Verify if expense in null
                        if myWS.cell(row=j, column=5).value != None:
                            exp = exp + myWS.cell(row=j, column=5).value
        myES.cell(row=i, column=maxCol).value = exp

# Main section of the program
# Get start and end dates for the workbook
startYear = int(input('Start Year: '))
startMonth = int(input('Start Month: '))
startDay = int(input('Start Day: '))
endYear = int(input('End Year: '))
endMonth = int(input('End Month: '))
endDay = int(input('End Day: '))
# Create start date and end dates
startDate = date(startYear, startMonth, startDay)
endDate = date(endYear, endMonth, endDay)
# Open Bank Accounts Excel file
wb = openpyxl.load_workbook(filePath + 'Bank Accounts.xlsx')
# Get worksheets in the Bank Accounts file
es = wb.get_sheet_by_name('Expenses')
maxCol = es.max_column + 1
# Get Major Bill worksheet and process expenses
fileName = 'Major Bills'
ws01 = wb.get_sheet_by_name(fileName)
calculateExpenses(ws01, es, startDate, endDate)
es.cell(row=2, column=maxCol).value = fileName
# Get Family worksheet and process expenses
maxCol += 1
fileName = 'Family'
ws02 = wb.get_sheet_by_name(fileName)
calculateExpenses(ws02, es, startDate, endDate)
es.cell(row=2, column=maxCol).value = fileName
maxCol += 1
fileName = 'Visa'
ws03 = wb.get_sheet_by_name(fileName)
calculateExpenses(ws03, es, startDate, endDate)
es.cell(row=2, column=maxCol).value = fileName
maxCol += 1
fileName = 'American Express'
ws04 = wb.get_sheet_by_name(fileName)
calculateExpenses(ws04, es, startDate, endDate)
es.cell(row=2, column=maxCol).value = fileName
maxCol += 1
fileName = 'BOA CC'
ws05 = wb.get_sheet_by_name(fileName)
calculateExpenses(ws05, es, startDate, endDate)
es.cell(row=2, column=maxCol).value = fileName
# Save the workbook
wb.save(filePath + 'Bank Accounts.xlsx')