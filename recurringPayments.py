import openpyxl
from datetime import date
from datetime import datetime

# Set variables
filePath = 'E:\\Excel\\'
startYear = int(input('Start Year: '))
startMonth = int(input('Start Month: '))
# Open Bank Accounts worksheet
wb = openpyxl.load_workbook(filePath + 'Bank Accounts.xlsx')
# Set Recurring Expense worksheet
reWS = wb.get_sheet_by_name('RecExp')
# Set Major Bills worksheet
mbWS = wb.get_sheet_by_name('Major Bills')
# Get max row in Major Bills worksheet
for j in range(2, 10000):
    if (mbWS.cell(row=j, column=1).value == None):
        maxRow = j
        break
# Loop through all of the recurring expenses
# on the RecExp worksheet
for i in range(2, 24):
    # Set due date variables
    dueDate = reWS.cell(row=i, column=4).value
    dueDay = int(dueDate.day)
    dueMonth = int(dueDate.month)
    dueYear = int(dueDate.year)
    # Determine if expense is due this month
    if (dueMonth == startMonth):
        # Get values from the RecExp worksheet
        freq = int(reWS.cell(row=i, column=5).value)
        # Add values to the Major Bills worksheet
        mbWS.cell(row=maxRow, column=1).value = reWS.cell(row=i, column=4).value
        mbWS.cell(row=maxRow, column=2).value = reWS.cell(row=i, column=1).value
        mbWS.cell(row=maxRow, column=5).value = reWS.cell(row=i, column=2).value
        mbWS.cell(row=maxRow, column=7).value = 'OBP'
        mbWS.cell(row=maxRow, column=9).value = 'N/A'
        mbWS.cell(row=maxRow, column=10).value = reWS.cell(row=i, column=3).value
        # Calculate new due date
        newMonth = dueMonth + freq
        if (newMonth > 12):
            newMonth = newMonth - 12
            dueYear = dueYear + 1
        newDate = date(dueYear, newMonth, dueDay)
        # Change due date on the RecExp worksheet
        reWS.cell(row=i, column=4).value = newDate
        # Calculate new maxRow
        maxRow = maxRow + 1
# Save the workbook
wb.save(filePath + 'Bank Accounts.xlsx')
